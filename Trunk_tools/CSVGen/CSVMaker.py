import openpyxl
import csv
import os
import time

# 允许的Unity数据类型列表
UNITY_DATATYPES = ["int", "float", "string", "bool", "Vector2", "Vector3", "Quaternion", "Color",'None']

# 读取Excel文件
def read_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    Table_data = []
    variable_names = []
    variable_datatypes = []

    # 获取变量名列表和参数类型列表（跳过第一个单元格）
    variable_names = [cell.value for cell in ws[4][1:]]
    variable_datatypes = [str(cell.value).lower() for cell in ws[3][1:]]  # 将数据类型转换为小写形式

    # 将数据类型转换为大写形式
    variable_datatypes = [datatype.capitalize() for datatype in variable_datatypes]

    # 检查第三行的数据类型是否正确填写
    invalid_datatypes = {}
    for idx, datatype in enumerate(variable_datatypes):
        if datatype.lower() not in [unity_type.lower() for unity_type in UNITY_DATATYPES]:
            invalid_datatypes[idx + 1] = datatype

    if invalid_datatypes:
        raise ValueError("Invalid datatypes specified in Excel sheet: {}".format(invalid_datatypes))

    # 遍历每一行（从第六行开始是实际数据）
    for row_idx in range(6, ws.max_row + 1):
        weapon_info = {}

        # 遍历每一列
        for col_idx in range(2, ws.max_column + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                weapon_info[variable_names[col_idx - 2]] = value

        Table_data.append(weapon_info)

    return Table_data, variable_datatypes

# 生成CSV文件
def generate_csv(csv_file_path, Table_data, variable_datatypes):
    # 写入校验的CSV文件
    with open(csv_file_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)

        # 写入变量名行
        writer.writerow(Table_data[0].keys())

        # 写入参数类型列表
        writer.writerow([datatype.lower() if datatype.lower() not in ['vector2', 'vector3'] else datatype.capitalize() for datatype in variable_datatypes])

        # 写入实际数据
        for item in Table_data:
            writer.writerow(item.values())



def main(input_folder_name, output_folder_name):
    print("=====================================================================================================")
    pattern=[
                "██╗   ██╗██╗   ██╗██╗   ██╗████████╗ ██████╗  ██████╗ ██╗     ███████╗",
                "╚██╗ ██╔╝██║   ██║██║   ██║╚══██╔══╝██╔═══██╗██╔═══██╗██║     ██╔════╝",
                " ╚████╔╝ ██║   ██║██║   ██║   ██║   ██║   ██║██║   ██║██║     ███████╗",
                "  ╚██╔╝  ██║   ██║██║   ██║   ██║   ██║   ██║██║   ██║██║     ╚════██║",
                "   ██║   ╚██████╔╝╚██████╔╝   ██║   ╚██████╔╝╚██████╔╝███████╗███████║",
                "   ╚═╝    ╚═════╝  ╚═════╝    ╚═╝    ╚═════╝  ╚═════╝ ╚══════╝╚══════╝",
                "                                                                      ",
                " ██████╗███████╗██╗   ██╗███╗   ███╗ █████╗ ██╗  ██╗███████╗██████╗   ",
                "██╔════╝██╔════╝██║   ██║████╗ ████║██╔══██╗██║ ██╔╝██╔════╝██╔══██╗  ",
                "██║     ███████╗██║   ██║██╔████╔██║███████║█████╔╝ █████╗  ██████╔╝  ",
                "██║     ╚════██║╚██╗ ██╔╝██║╚██╔╝██║██╔══██║██╔═██╗ ██╔══╝  ██╔══██╗  ",
                "╚██████╗███████║ ╚████╔╝ ██║ ╚═╝ ██║██║  ██║██║  ██╗███████╗██║  ██║  ",
                " ╚═════╝╚══════╝  ╚═══╝  ╚═╝     ╚═╝╚═╝  ╚═╝╚═╝  ╚═╝╚══════╝╚═╝  ╚═╝  ",                                  
            ]
    for line in pattern:
        print(line)
    print("=====================================================================================================")
    time.sleep(1)
    start_time = time.time()  # 记录开始时间

    # 清空输出路径中的CSV文件
    clear_output_folder(output_folder_name)

    # 遍历输入文件夹中的所有xlsx文件
    num_processed_files = 0
    num_Error_files = 0
    invalid_datatypes = {}  # 记录不合规的数据类型
    for filename in os.listdir(input_folder_name):
        if filename.endswith('.xlsx'):
            excel_file_path = os.path.join(input_folder_name, filename)
            csv_file_path = os.path.join(output_folder_name, filename[:-5] + '.csv')  # 构建对应的输出路径

            try:
                # 读取Excel文件
                weapon_data, variable_datatypes = read_excel(excel_file_path)

                # 生成CSV文件
                generate_csv(csv_file_path, weapon_data, variable_datatypes)

                # 打印成功信息
                print()
                print("=====================================================================================================")
                print("Successfully generated CSV:", filename[:-5] + '.csv')
                print("Total items:", len(weapon_data))
                print("Items information:", len(variable_datatypes))
                print("=====================================================================================================")
                print()
                print()
                num_processed_files += 1

            except ValueError as e:
                invalid_datatypes[filename] = str(e)

    end_time = time.time()  # 记录结束时间
    total_time = end_time - start_time  # 计算总耗时

    # 打印警告信息
    if invalid_datatypes:
        print()
        print("=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=")
        for filename, warning in invalid_datatypes.items():
            print("Error DataType Found In CSV:", filename[:-5] + '.csv')
            print("Invalid datatypes in columns:")
            print(warning)
        print("=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=")
        print()
        print()
        num_Error_files +=1


    if(num_Error_files != 0):
                pattern = [
                            "=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=",
                            "██╗  ██╗      ██████╗ ██╗   ██╗██╗██╗     ██████╗     ███████╗ █████╗ ██╗██╗     ██╗   ██╗██████╗ ███████╗   ██╗  ██╗",
                            "╚██╗██╔╝      ██╔══██╗██║   ██║██║██║     ██╔══██╗    ██╔════╝██╔══██╗██║██║     ██║   ██║██╔══██╗██╔════╝   ╚██╗██╔╝",
                            " ╚███╔╝ █████╗██████╔╝██║   ██║██║██║     ██║  ██║    █████╗  ███████║██║██║     ██║   ██║██████╔╝█████╗█████╗╚███╔╝ ",
                            " ██╔██╗ ╚════╝██╔══██╗██║   ██║██║██║     ██║  ██║    ██╔══╝  ██╔══██║██║██║     ██║   ██║██╔══██╗██╔══╝╚════╝██╔██╗ ",
                            "██╔╝ ██╗      ██████╔╝╚██████╔╝██║███████╗██████╔╝    ██║     ██║  ██║██║███████╗╚██████╔╝██║  ██║███████╗   ██╔╝ ██╗",
                            "╚═╝  ╚═╝      ╚═════╝  ╚═════╝ ╚═╝╚══════╝╚═════╝     ╚═╝     ╚═╝  ╚═╝╚═╝╚══════╝ ╚═════╝ ╚═╝  ╚═╝╚══════╝   ╚═╝  ╚═╝",
                            "=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!="
                            ]
                for line in pattern:
                    print(line)
                print()
                print("=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=")
                print("Total time elapsed: {:.2f} seconds".format(total_time))
                print("Number of files processed:", num_processed_files)
                print("Some Tables' DataType Invalid,they will be Ignored while Generating")
                print("Number of files Skipped:", num_Error_files)
                print("=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=!=")
                print()
                print()
    else:
                pattern = [
                            "===================================================================================================================",
                            "██╗      ██████╗ ██╗   ██╗██╗██╗     ██████╗     ███████╗██╗   ██╗ ██████╗ ██████╗███████╗███████╗███████╗      ██╗",
                            "██║      ██╔══██╗██║   ██║██║██║     ██╔══██╗    ██╔════╝██║   ██║██╔════╝██╔════╝██╔════╝██╔════╝██╔════╝      ██║",
                            "██║█████╗██████╔╝██║   ██║██║██║     ██║  ██║    ███████╗██║   ██║██║     ██║     █████╗  ███████╗███████╗█████╗██║",
                            "╚═╝╚════╝██╔══██╗██║   ██║██║██║     ██║  ██║    ╚════██║██║   ██║██║     ██║     ██╔══╝  ╚════██║╚════██║╚════╝╚═╝",
                            "██╗      ██████╔╝╚██████╔╝██║███████╗██████╔╝    ███████║╚██████╔╝╚██████╗╚██████╗███████╗███████║███████║      ██╗",
                            "╚═╝      ╚═════╝  ╚═════╝ ╚═╝╚══════╝╚═════╝     ╚══════╝ ╚═════╝  ╚═════╝ ╚═════╝╚══════╝╚══════╝╚══════╝      ╚═╝",
                            "==================================================================================================================="
                            ]
                for line in pattern:
                    print(line)
                    # 打印耗时和处理的文件数量
                print()
                print("===================================================================================================================",)
                print("Total time elapsed: {:.2f} seconds".format(total_time))
                print("Number of files processed:", num_processed_files)
                print("All Tables Passed DataTypeCheck!")
                print("===================================================================================================================",)
                print()
                print()

def clear_output_folder(output_folder_path):
    for filename in os.listdir(output_folder_path):
        file_path = os.path.join(output_folder_path, filename)
        os.remove(file_path)
    print()
    print("-------------------------------------------------------------------------------------------------------------------")
    print("Output folder cleared.")
    print("-------------------------------------------------------------------------------------------------------------------")
    print()

if __name__ == "__main__":
    # 获取当前脚本文件所在目录
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # 构建输入文件夹的相对路径
    input_folder_name = os.path.normpath(os.path.join(script_dir, '../Table'))

    # 构建输出文件夹的相对路径
    output_folder_name = os.path.normpath(os.path.join(script_dir, '../../Assets/Resources/Data/Table'))

    main(input_folder_name, output_folder_name)
